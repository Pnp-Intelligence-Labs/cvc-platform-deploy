"""
scorecard.py — Generate a scoring Excel workbook for a DD run.

Reads workdir/[company]/overview.json + agents/*.json and writes:
    workdir/[company]/[company]_Scorecard.xlsx

Sheets:
    Summary         — one-page snapshot: recommendation, growth metrics status, tech score, flags
    Growth Metrics  — key metrics vs model-specific benchmarks (auto-scored from dataroom)
    Tech Score      — 9-section rubric with points math (auto-scored from Q&A; shows rubric if Q&A absent)
    Checklist       — preliminary diligence checklist coverage
    All Findings    — every finding across all agents

Run:
    python3 scorecard.py "Retina Robotics"
    python3 scorecard.py "Retina Robotics" --output /some/path
"""

import json
import argparse
from pathlib import Path
from datetime import datetime

from config.settings import WORKDIR

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── Brand colours ──────────────────────────────────────────────────────────────
NAVY   = "253B49"
YELLOW = "F0E545"
BG     = "F5F5F7"
WHITE  = "FFFFFF"
RED    = "C0392B"
GREEN  = "27AE60"
ORANGE = "E67E22"
GREY   = "676E7A"

AGENTS = ["financials", "comp", "qualitative", "product", "news"]

# ── Tech rubric reference (for when Q&A is not provided) ──────────────────────
# (section_key, display_name, max_possible, [criteria strings])
TECH_RUBRIC = [
    ("code_history", "Code History", 12, [
        "All software coded in-house (excl. open source/APIs): +3 | non-core outsourced: +1 | core outsourced: -1 | all outsourced: -10",
        "Initial developer is still main developer: 0 | no longer main developer: -5",
        "Built similar product before: +2",
        "No single points of failure: +2 | known only by dev WITH equity: -1 | WITHOUT equity: -2",
    ]),
    ("agility", "Agility", 10, [
        "Continuous deployment: +5 | monthly releases: 0 | 5x/year: -3 | 1x/year: -5",
        "Build vs buy: evaluates build AND maintain hours: +2 | build hours only: -1 | builds first: -5 | buys first: 0",
        "Payment system: 3rd party (Stripe, etc.): 0 | custom in-house: -10",
        "Clients paying outside payment system: -2",
        "Invoicing: 3rd party: 0 | custom in-house: -5",
    ]),
    ("monitoring", "Monitoring", 8, [
        "Application performance monitored: +1",
        "Application security monitored: +1",
        "Infrastructure performance monitored: +1",
        "Website monitoring: +1",
        "Exception monitoring: +1",
        "Internally developed monitoring (vs 3rd party): -2",
        "Max system capacity measured: +2",
    ]),
    ("compliance", "Compliance & Security", 11, [
        "3rd party data: licensed: 0 | crawled non-critical: -1 | crawled critical: -2",
        "Library/open-source license monitoring: +1 | server/desktop only: 0 | not monitored: -2",
        "Version control backed up to 3rd party: +2",
        "Database: 0% data loss possible: +5 | 50% loss: -5 | 100% loss: -10",
        "Disaster recovery plan exists: +1",
    ]),
    ("processes", "Product Dev & Processes", 22, [
        "Version control: cloud hosted (GitHub): +2 | local only: +1 | none: -20",
        "Unit tests: critical routines only: +5 | 100% coverage: -1 | none: -5",
        "Code reviews: critical routines only: +5 | all code: -1 | none: -5",
        "One-click deploy to staging/production: +2 | no: -5",
        "Feature flags system: +2 | no: -3",
        "Show features to limited users without hardcoding: +2 | no: -3",
        "Hosting: mix of IaaS: +1 | single IaaS: 0 | PaaS only: -1",
        "Heavy cron jobs: 0 | has cron jobs: -2",
        "Queueing system between jobs: +3 | no queueing: -3",
        "3rd party providers with less funding: none: 0 | some: -1 | all: -5",
    ]),
    ("tech_org", "Tech Organization", 19, [
        "Roadmap owned by CPO: +5 | CEO: +2 | CTO: -5",
        "CTO speaks with customers weekly (calls): +5 | weekly (tickets): +2 | never: -5",
        "Tech team speaks with customers weekly: +2 | CTO only: -2 | never: -5",
        "Power users available for feedback: +2",
        "Roadmap visible to customers: +2 | all internally: 0 | tech team only: -3 | leadership only: -4 | in founder's head: -5",
    ]),
    ("leadership", "Tech Founder Leadership", 18, [
        "Founder can convincingly pitch: +5 | cannot: -5",
        "Last customer conversation: this week: +3 | this month: +1 | >1 month: -1 | never: -10",
        "Product roadmap written 6 months ahead: +2 | 1 month: +1 | >12 months: +1 | no written plan: -5",
        "Engineering values written down: +2",
        "Application performance technology prioritized: +1 | reliability: +1",
        "Who decides tech stack: team with founder final say: +1 | team consensus: 0 | devs decide: -1 | founder alone: -5",
        "Scaling awareness (x10/x100/x10000): aware of bottlenecks: +2 | solves when appears: -2 | scale horizontally: -5 | doesn't know: -5",
    ]),
    ("hiring", "Hiring", 18, [
        "All developers worked for founder before: +3 | most: +2 | some: +1 | none: -1",
        "All developers have equity: +3 | most: +2 | some: +1",
        "All new hires from referrals: +3 | most: +2 | some: +1 | none: -1",
        "Interview includes team interview: +1 | coding during interview: +1",
        "Reference calls with backdoor references: +5 | provided refs only: 0 | never: -10",
        "Dev vacancy applications: 50+: +2 | 11-50: +1 | 0-10: 0",
        "Time to hire: weeks or 1 month: +1 | 1 week or months: -2",
    ]),
    ("people_mgmt", "People Management", 8, [
        "Team attrition last year: none: 0 | some: -5 | all: -10",
        "Reason for leaving: lack of motivation: -5 | salary: -1 | other: 0",
        "1:1 frequency: daily: +3 | weekly: +2 | monthly: -2 | never: -10",
    ]),
]

TECH_MAX_TOTAL = sum(s[2] for s in TECH_RUBRIC)  # 126


# ── Growth metric benchmarks ───────────────────────────────────────────────────

def _growth_metric_status(key: str, value, model_type: str):
    """Return (status, benchmark_label) for a key metric."""
    if value is None:
        return "na", "—"

    try:
        value = float(value)
    except (TypeError, ValueError):
        return "na", "—"

    model_type = (model_type or "").lower()

    if key == "runway":
        # Moonshots need more runway to hit milestones
        if model_type == "moonshot_hardtech":
            if value >= 24: return "pass", "≥24 months"
            if value >= 18: return "warn", "≥24 months"
            return "fail", "≥24 months"
        else:
            if value >= 18: return "pass", "≥18 months"
            if value >= 12: return "warn", "≥18 months"
            return "fail", "≥18 months"

    if key == "burn_multiple":
        if value <= 1.5: return "pass", "<1.5x"
        if value <= 3.0: return "warn", "<1.5x"
        return "fail", "<1.5x"

    if key == "gross_margin":
        if model_type in ("saas", "enterprise_saas", "usage_based"):
            thr = 60
        elif model_type == "marketplace":
            thr = 50
        elif model_type in ("hardware", "moonshot_hardtech"):
            thr = 30
        elif model_type == "ecommerce":
            thr = 40
        else:
            thr = 50
        label = f"≥{thr}%"
        if value >= thr:       return "pass", label
        if value >= thr * 0.7: return "warn", label
        return "fail", label

    if key == "revenue_growth":
        if model_type in ("saas", "enterprise_saas", "usage_based"):
            thr, label = 200, ">200% YoY"
        elif model_type in ("marketplace", "ecommerce", "hardware"):
            thr, label = 20, ">20% MoM"
        elif model_type == "moonshot_hardtech":
            # Moonshot: no growth rate benchmark — milestone-based
            return "na", "Milestone-based (no growth rate benchmark)"
        else:
            thr, label = 100, ">100% YoY"
        if value >= thr:       return "pass", label
        if value >= thr * 0.5: return "warn", label
        return "fail", label

    if key == "arr":
        # ARR is informational — no universal pass/fail threshold
        return "info", "—"

    if key == "burn_rate":
        # Burn is informational in context of runway
        return "info", "See Runway"

    return "na", "—"


def _fmt_metric(key: str, value) -> str:
    if value is None:
        return "N/A"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    if key == "arr":
        return f"${v:,.0f}/yr" if v >= 1 else "$0"
    if key == "burn_rate":
        return f"${v:,.0f}/mo"
    if key == "runway":
        return f"{v:.1f} months"
    if key == "gross_margin":
        return f"{v:.0f}%"
    if key == "burn_multiple":
        return f"{v:.2f}x"
    if key == "revenue_growth":
        return f"{v:.0f}%"
    return str(value)


STATUS_COLOR = {
    "pass": GREEN,
    "warn": ORANGE,
    "fail": RED,
    "na":   GREY,
    "info": GREY,
}

STATUS_LABEL = {
    "pass": "PASS",
    "warn": "WATCH",
    "fail": "FAIL",
    "na":   "N/A",
    "info": "—",
}

GROWTH_METRIC_ROWS = [
    ("ARR / MRR",      "arr"),
    ("Revenue Growth", "revenue_growth"),
    ("Gross Margin",   "gross_margin"),
    ("Burn Rate",      "burn_rate"),
    ("Runway",         "runway"),
    ("Burn Multiple",  "burn_multiple"),
]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=None, size=10) -> Font:
    return Font(bold=bold, color=color or "000000", size=size, name="Calibri")

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def _border() -> Border:
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _header_row(ws, row: int, values: list, bg=NAVY, fg=WHITE, height=20):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, color=fg, size=10)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = height

def _data_row(ws, row: int, values: list, bg=WHITE, bold=False, height=15):
    for col, val in enumerate(values, 1):
        if isinstance(val, list):
            val = ", ".join(str(v) for v in val)
        elif isinstance(val, dict):
            val = str(val)
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(bg)
        c.font      = _font(bold=bold)
        c.alignment = _left()
        c.border    = _border()
    ws.row_dimensions[row].height = height

def _score_color(score) -> str:
    if score is None: return GREY
    if score > 0:     return GREEN
    if score < 0:     return RED
    return "888888"

def _score_grade_pct(pct: int) -> str:
    if pct >= 80: return "A"
    if pct >= 65: return "B"
    if pct >= 50: return "C"
    if pct >= 35: return "D"
    return "F"


# ── Load data ──────────────────────────────────────────────────────────────────

def load_data(company: str, version: str = None) -> dict:
    safe      = company.replace(" ", "_").replace("/", "-")
    base_dir  = WORKDIR / safe
    suffix    = f"_{version}" if version else ""
    agents_subdir = "agents_v2" if version == "v2" else "agents"

    overview   = {}
    agent_data = {}

    overview_path = base_dir / f"overview{suffix}.json"
    if overview_path.exists():
        overview = json.loads(overview_path.read_text())

    for agent in AGENTS:
        path = base_dir / agents_subdir / f"{agent}.json"
        if path.exists():
            agent_data[agent] = json.loads(path.read_text())

    return {"company": company, "overview": overview, "agents": agent_data}


# ── Sheet 1: Summary ───────────────────────────────────────────────────────────

def build_summary_sheet(ws, data: dict):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    company  = data["company"]
    overview = data["overview"]
    sc       = overview.get("scorecard", {})
    km       = overview.get("key_metrics", {})
    model    = sc.get("business_model", "")

    row = 1

    # Title
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value=f"DD Scorecard — {company}")
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=WHITE, size=14)
    c.alignment = _center(); ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d')}")
    c.fill = _fill(BG); c.font = _font(color=GREY, size=9); c.alignment = _center()
    row += 2

    def section(title):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.fill = _fill(YELLOW); c.font = _font(bold=True, color=NAVY, size=10)
        c.alignment = _left(); ws.row_dimensions[row].height = 16
        row += 1

    def kv(key, val, bold_val=False):
        nonlocal row
        _data_row(ws, row, [key, val], bg=WHITE if row % 2 == 0 else BG)
        if bold_val:
            ws.cell(row=row, column=2).font = _font(bold=True)
        row += 1

    # Recommendation
    section("RECOMMENDATION")
    rec = overview.get("recommendation", "—").upper()
    rec_color = {"STRONG_INTEREST": GREEN, "PROCEED": GREEN,
                 "CONDITIONAL": ORANGE, "PASS": RED}.get(rec, GREY)
    kv("Decision", rec, bold_val=True)
    ws.cell(row=row-1, column=2).font = _font(bold=True, color=rec_color, size=11)
    kv("Rationale", overview.get("recommendation_rationale", "—"))
    kv("Stage",         overview.get("stage", "—"))
    kv("Raise Amount",  overview.get("raise_amount", "—"))
    kv("Valuation Ask", overview.get("valuation_ask", "—"))
    row += 1

    # Growth Metrics Status
    section("GROWTH METRICS  (see Growth Metrics sheet for detail)")
    kv("Business Model",   model.replace("_", " ").title() if model else "—")
    kv("Growth Benchmark", sc.get("growth_benchmark", "—"))
    fin = data["agents"].get("financials", {})
    fs  = fin.get("financial_score", {})
    if fs.get("total") is not None:
        pct = round(fs["total"] / fs.get("max_possible", 17) * 100)
        kv("Financial Score", f"{fs['total']} / {fs.get('max_possible', 17)}  ({pct}%)")
        ws.cell(row=row-1, column=2).font = _font(bold=True, color=_score_color(fs["total"]))
    else:
        kv("Financial Score", "Pending re-run with scoring prompt")
        ws.cell(row=row-1, column=2).font = _font(color=ORANGE)
    row += 1

    # Tech Score
    section("TECH SCORE  (see Tech Score sheet for detail)")
    ts = sc.get("tech_score", {})
    if ts.get("total") is not None:
        pct = round(ts["total"] / TECH_MAX_TOTAL * 100)
        grade = _score_grade_pct(pct)
        kv("Score",   f"{ts['total']} / {TECH_MAX_TOTAL}")
        kv("Percent", f"{pct}%")
        kv("Grade",   grade, bold_val=True)
        grade_colors = {"A": GREEN, "B": GREEN, "C": ORANGE, "D": RED, "F": RED}
        ws.cell(row=row-1, column=2).font = _font(bold=True, color=grade_colors.get(grade, GREY), size=11)
        ss = ts.get("section_scores", {})
        for sect_key, s in ss.items():
            s_score = s.get("score", 0)
            s_max   = s.get("max_possible", 0)
            label   = sect_key.replace("_", " ").title()
            kv(f"  {label}", f"{s_score} / {s_max}")
            ws.cell(row=row-1, column=2).font = _font(bold=True, color=_score_color(s_score))
    else:
        kv("Status", "Q&A not provided — see Tech Score sheet for required questions")
        ws.cell(row=row-1, column=2).font = _font(color=ORANGE)
    row += 1

    # Checklist Coverage
    section("DILIGENCE CHECKLIST COVERAGE")
    cc = sc.get("checklist_coverage", {})
    if cc:
        for sname, status in cc.items():
            color = GREEN if status == "complete" else (ORANGE if status == "partial" else RED)
            kv(sname.replace("_", " ").title(), status)
            ws.cell(row=row-1, column=2).font = _font(bold=True, color=color)
    else:
        kv("Coverage", "Not available")
    row += 1

    # Flags
    section("FLAGS SUMMARY")
    all_flags    = overview.get("all_flags", [])
    red_flags    = [f for f in all_flags if f.get("severity") == "red"]
    yellow_flags = [f for f in all_flags if f.get("severity") == "yellow"]
    kv("Red Flags",    str(len(red_flags)))
    ws.cell(row=row-1, column=2).font = _font(bold=True, color=RED if red_flags else GREY)
    kv("Yellow Flags", str(len(yellow_flags)))
    ws.cell(row=row-1, column=2).font = _font(bold=True, color=ORANGE if yellow_flags else GREY)


# ── Sheet 2: Growth Metrics ────────────────────────────────────────────────────

def build_growth_metrics_sheet(ws, data: dict):
    ws.title = "Growth Metrics"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22   # Topic
    ws.column_dimensions["B"].width = 22   # Actual
    ws.column_dimensions["C"].width = 8    # Score
    ws.column_dimensions["D"].width = 8    # Max
    ws.column_dimensions["E"].width = 10   # Flag
    ws.column_dimensions["F"].width = 42   # Finding / Score Reason

    overview = data["overview"]
    sc       = overview.get("scorecard", {})
    km       = overview.get("key_metrics", {})
    model    = sc.get("business_model", "")
    fin      = data["agents"].get("financials", {})
    fs       = fin.get("financial_score", {})

    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value=f"Growth Metrics — {data['company']}")
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=WHITE, size=12)
    c.alignment = _center(); ws.row_dimensions[row].height = 24
    row += 2

    # Model + benchmark + total score header
    total  = fs.get("total")
    max_fs = fs.get("max_possible", 17)
    score_str = f"Financial Score: {total} / {max_fs}" if total is not None else "Financial Score: Pending re-run"
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1,
                value=f"Business Model: {model.replace('_', ' ').title() if model else '—'}    |    "
                      f"Benchmark: {sc.get('growth_benchmark', '—')}    |    {score_str}")
    c.fill = _fill(YELLOW); c.font = _font(bold=True, color=NAVY)
    c.alignment = _left(); ws.row_dimensions[row].height = 16
    row += 1

    # Section scores if available
    section_scores = fs.get("section_scores", {})
    if section_scores:
        ws.merge_cells(f"A{row}:F{row}")
        parts = [f"{k.replace('_', ' ').title()}: {v}" for k, v in section_scores.items()]
        c = ws.cell(row=row, column=1, value="  |  ".join(parts))
        c.fill = _fill(BG); c.font = _font(color=GREY, size=9); c.alignment = _left()
        ws.row_dimensions[row].height = 14
        row += 1

    row += 1

    # Findings-based rows (scored by agent)
    findings_by_topic = {}
    for f in fin.get("findings", []):
        if isinstance(f, dict) and f.get("topic") not in ("missing_document",):
            findings_by_topic[f.get("topic")] = f

    if findings_by_topic:
        _header_row(ws, row, ["Topic", "Actual", "Score", "Max", "Flag", "Finding & Score Reason"])
        row += 1

        TOPIC_MAX = {
            "arr": 2, "revenue_growth": 2, "burn_rate": 1, "runway": 2,
            "gross_margin": 2, "burn_multiple": 2, "revenue_concentration": 1,
            "customer_contracts": 2, "cap_table": 1, "valuation": 1,
            "path_to_profitability": 1,
        }

        for topic, max_pts in TOPIC_MAX.items():
            f = findings_by_topic.get(topic)
            if not f:
                continue
            score     = f.get("score")
            flag      = f.get("flag", False)
            actual    = _fmt_metric(topic, km.get(topic)) if topic in km else (f.get("our_finding", "—") or "—")
            # Truncate long finding text
            finding_text = str(f.get("our_finding", "") or "")
            score_reason = str(f.get("score_reason", "") or "")
            cell_text = score_reason if score_reason and score_reason != "None" else finding_text
            if len(cell_text) > 140:
                cell_text = cell_text[:140] + "..."

            bg = "FFF3F3" if flag else (BG if row % 2 == 0 else WHITE)
            _data_row(ws, row, [
                topic.replace("_", " ").title(),
                actual,
                score if score is not None else "—",
                max_pts,
                "YES" if flag else "",
                cell_text,
            ], bg=bg, height=28)
            if score is not None:
                ws.cell(row=row, column=3).font = _font(bold=True, color=_score_color(score))
            if flag:
                ws.cell(row=row, column=5).font = _font(bold=True, color=RED)
            row += 1

        # Total row
        if total is not None:
            _data_row(ws, row, ["TOTAL", "", total, max_fs, "", ""], bg=BG, bold=True)
            ws.cell(row=row, column=3).font = _font(bold=True, color=_score_color(total), size=11)
            row += 1

    else:
        # Agent hasn't run yet — show key metrics with scorecard thresholds as fallback
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1,
                    value="Financials agent output does not contain scores. Re-run pipeline to generate scored findings.")
        c.fill = _fill("FFF3CD"); c.font = _font(bold=True, color="856404"); c.alignment = _left()
        ws.row_dimensions[row].height = 20
        row += 2

        _header_row(ws, row, ["Metric", "Actual", "—", "—", "—", "Notes"])
        row += 1
        for label, key in GROWTH_METRIC_ROWS:
            val = km.get(key)
            display = _fmt_metric(key, val)
            notes = _metric_notes(key, data)
            _data_row(ws, row, [label, display, "—", "—", "—", notes],
                      bg=BG if row % 2 == 0 else WHITE)
            row += 1

    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1,
                value="Scores produced by Financials Agent reconciliation (FINANCIALS_RECONCILE). "
                      "Max possible: 17. Sections: Growth (4), Efficiency (3), Durability (2), "
                      "Unit Economics (3), Deal Structure (4), Path to Profit (1).")
    c.fill = _fill(BG); c.font = _font(color=GREY, size=9); c.alignment = _left()
    ws.row_dimensions[row].height = 24


def _metric_notes(key: str, data: dict) -> str:
    """Pull a brief note from financials findings for this metric."""
    fin = data["agents"].get("financials", {})
    findings = fin.get("findings", [])
    for f in findings:
        if not isinstance(f, dict):
            continue
        if f.get("topic") == key:
            finding = str(f.get("our_finding", "") or "")
            return finding[:120] + "..." if len(finding) > 120 else finding
    return "—"


# ── Sheet 3: Tech Score ────────────────────────────────────────────────────────

def build_tech_sheet(ws, data: dict):
    ws.title = "Tech Score"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    overview = data["overview"]
    sc       = overview.get("scorecard", {})
    ts       = sc.get("tech_score", {})
    agents   = data["agents"]
    product  = agents.get("product", {})

    tech_findings = [f for f in product.get("findings", [])
                     if isinstance(f, dict) and f.get("id", "").startswith("product_tech_")]

    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value=f"Tech Score — {data['company']}")
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=WHITE, size=12)
    c.alignment = _center(); ws.row_dimensions[row].height = 24
    row += 2

    has_score = ts.get("total") is not None
    has_section_scores = bool(ts.get("section_scores"))

    if has_score or has_section_scores:
        # ── Section summary ────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="SECTION SCORES")
        c.fill = _fill(YELLOW); c.font = _font(bold=True, color=NAVY)
        c.alignment = _left(); row += 1

        _header_row(ws, row, ["Section", "Score", "Max", "Pct", "Grade", ""])
        row += 1

        section_scores = ts.get("section_scores", {})
        for sect_key, display_name, max_pts, _ in TECH_RUBRIC:
            s = section_scores.get(sect_key, {})
            score = s.get("score")
            maxi  = s.get("max_possible", max_pts)
            if score is None:
                _data_row(ws, row, [display_name, "—", maxi, "—", "—", ""], bg=BG if row % 2 == 0 else WHITE)
            else:
                pct   = round(score / maxi * 100) if maxi else 0
                grade = _score_grade_pct(pct)
                bg    = "FFF3F3" if score < 0 else (BG if row % 2 == 0 else WHITE)
                _data_row(ws, row, [display_name, score, maxi, f"{pct}%", grade, ""], bg=bg)
                ws.cell(row=row, column=2).font = _font(bold=True, color=_score_color(score))
            row += 1

        if has_score:
            total = ts["total"]
            pct   = round(total / TECH_MAX_TOTAL * 100)
            grade = _score_grade_pct(pct)
            _data_row(ws, row, ["TOTAL", total, TECH_MAX_TOTAL, f"{pct}%", grade, ""],
                      bg=BG, bold=True)
            ws.cell(row=row, column=2).font = _font(bold=True, color=_score_color(total), size=11)
            ws.cell(row=row, column=5).font = _font(bold=True, size=11)
            row += 1

        row += 1

        if tech_findings:
            # ── Individual findings ────────────────────────────────────────────
            ws.merge_cells(f"A{row}:F{row}")
            c = ws.cell(row=row, column=1, value="INDIVIDUAL FINDINGS — from Q&A")
            c.fill = _fill(YELLOW); c.font = _font(bold=True, color=NAVY)
            c.alignment = _left(); row += 1

            _header_row(ws, row, ["Section", "Founder Answer", "Our Assessment", "Points", "Max", "Flag"])
            row += 1

            # Group findings by section for readability
            section_order = [s[0] for s in TECH_RUBRIC]
            findings_by_section: dict = {}
            for f in tech_findings:
                sec = f.get("topic", "other")
                findings_by_section.setdefault(sec, []).append(f)

            for sect_key in section_order:
                sect_findings = findings_by_section.get(sect_key, [])
                if not sect_findings:
                    continue
                # Section label row
                ws.merge_cells(f"A{row}:F{row}")
                c = ws.cell(row=row, column=1,
                            value=next((s[1] for s in TECH_RUBRIC if s[0] == sect_key), sect_key.title()))
                c.fill = _fill(BG); c.font = _font(bold=True, color=NAVY, size=9)
                c.alignment = _left(); ws.row_dimensions[row].height = 14
                row += 1
                for f in sect_findings:
                    score = f.get("score")
                    flag  = f.get("flag", False)
                    bg    = "FFF3F3" if flag else (WHITE if row % 2 == 0 else BG)
                    # Max for individual finding: not stored, so leave blank
                    _data_row(ws, row, [
                        f.get("topic", "").replace("_", " ").title(),
                        f.get("claimed") or "—",
                        f.get("our_finding") or "—",
                        score if score is not None else "—",
                        "—",
                        "YES" if flag else "",
                    ], bg=bg, height=30)
                    if score is not None:
                        ws.cell(row=row, column=4).font = _font(bold=True, color=_score_color(score))
                    if flag:
                        ws.cell(row=row, column=6).font = _font(bold=True, color=RED)
                    row += 1

    else:
        # ── No Q&A provided: show rubric as reference ──────────────────────────
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1,
                    value="Q&A NOT PROVIDED — Tech score requires investor questionnaire responses. "
                          "Rubric below shows what will be scored and max points per section.")
        c.fill = _fill("FFF3CD"); c.font = _font(bold=True, color="856404")
        c.alignment = _left(); ws.row_dimensions[row].height = 28
        row += 2

        _header_row(ws, row, ["Section", "Max Points", "% of Total", "Status", "Scoring Criteria (abbreviated)", ""])
        row += 1

        for sect_key, display_name, max_pts, criteria in TECH_RUBRIC:
            pct_of_total = f"{round(max_pts / TECH_MAX_TOTAL * 100)}%"
            bg = BG if row % 2 == 0 else WHITE

            # First row: section name, max, pct, status
            _data_row(ws, row, [display_name, max_pts, pct_of_total, "PENDING", "", ""], bg=bg)
            ws.cell(row=row, column=4).font = _font(bold=True, color=ORANGE)
            ws.cell(row=row, column=2).font = _font(bold=True)
            row += 1

            # Criteria rows (indented)
            for criterion in criteria:
                _data_row(ws, row, ["", "", "", "", criterion, ""],
                          bg=bg, height=24)
                ws.cell(row=row, column=5).font = _font(color=GREY, size=9)
                row += 1

        row += 1

        # Total row
        _data_row(ws, row, ["TOTAL MAX", TECH_MAX_TOTAL, "100%", "—", "", ""], bg=BG, bold=True)
        row += 2

        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1,
                    value="To score: add an investor Q&A document to the dataroom and re-run the pipeline. "
                          "The Q&A should address the categories above — particularly Code History, "
                          "Processes, and Tech Org, which carry the most weight.")
        c.fill = _fill(BG); c.font = _font(color=GREY, size=9); c.alignment = _left()
        ws.row_dimensions[row].height = 36


# ── Sheet 4: Checklist ─────────────────────────────────────────────────────────

def build_checklist_sheet(ws, data: dict):
    ws.title = "Checklist"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14

    agents = data["agents"]
    qual   = agents.get("qualitative", {})

    org_findings = [f for f in qual.get("findings", [])
                    if isinstance(f, dict) and f.get("id", "").startswith("qualitative_org_")]

    row = 1
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"Preliminary Diligence Checklist — {data['company']}")
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=WHITE, size=12)
    c.alignment = _center(); ws.row_dimensions[row].height = 24
    row += 2

    sc = data["overview"].get("scorecard", {})
    cc = sc.get("checklist_coverage", {})
    if cc:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value="COVERAGE SUMMARY")
        c.fill = _fill(YELLOW); c.font = _font(bold=True, color=NAVY); c.alignment = _left()
        row += 1
        for sname, status in cc.items():
            color = GREEN if status == "complete" else (ORANGE if status == "partial" else RED)
            _data_row(ws, row, [sname.replace("_", " ").title(), status, "", ""])
            ws.cell(row=row, column=2).font = _font(bold=True, color=color)
            row += 1
        row += 1

    if org_findings:
        _header_row(ws, row, ["Topic", "Answer Found", "Verdict", "Flagged"])
        row += 1
        for f in org_findings:
            bg = "FFF3F3" if f.get("flag") else (BG if row % 2 == 0 else WHITE)
            _data_row(ws, row, [
                f.get("topic", "").replace("_", " ").title(),
                f.get("claimed") or "—",
                f.get("verdict", "—"),
                "YES" if f.get("flag") else "",
            ], bg=bg)
            if f.get("flag"):
                ws.cell(row=row, column=4).font = _font(bold=True, color=RED)
            row += 1
    else:
        _data_row(ws, row, ["No checklist findings available", "", "", ""])


# ── Sheet 5: All Findings ──────────────────────────────────────────────────────

def build_findings_sheet(ws, data: dict):
    ws.title = "All Findings"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 20   # Accuracy (feedback)
    ws.column_dimensions["H"].width = 22   # Flag Rating (feedback)
    ws.column_dimensions["I"].width = 40   # Notes (feedback)

    row = 1
    ws.merge_cells(f"A{row}:I{row}")
    c = ws.cell(row=row, column=1, value=f"All Findings — {data['company']}")
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=WHITE, size=12)
    c.alignment = _center(); ws.row_dimensions[row].height = 24
    row += 1

    # Feedback instructions banner
    ws.merge_cells(f"A{row}:I{row}")
    c = ws.cell(row=row, column=1,
                value="FEEDBACK COLUMNS (G-I): Use dropdowns in Accuracy and Flag Rating to rate each finding. "
                      "Add corrections or missed context in Notes. This data is used to improve the pipeline.")
    c.fill = _fill("FFF3CD"); c.font = _font(color="856404", size=9); c.alignment = _left()
    ws.row_dimensions[row].height = 18
    row += 2

    _header_row(ws, row, [
        "Agent", "Topic", "Our Finding", "Claimed", "Verdict", "Flag",
        "Accuracy", "Flag Rating", "Notes / Correction",
    ])
    # Highlight feedback headers in yellow
    for col in (7, 8, 9):
        ws.cell(row=row, column=col).fill = _fill(YELLOW)
        ws.cell(row=row, column=col).font = _font(bold=True, color=NAVY, size=10)
    header_row = row
    row += 1

    data_start_row = row

    for agent in AGENTS:
        agent_out = data["agents"].get(agent, {})
        findings  = agent_out.get("findings", [])
        for f in findings:
            if not isinstance(f, dict):
                continue
            flag = f.get("flag", False)
            bg   = "FFF3F3" if flag else (BG if row % 2 == 0 else WHITE)
            _data_row(ws, row, [
                agent.title(),
                f.get("topic", "").replace("_", " "),
                f.get("our_finding", "—"),
                f.get("claimed") or "—",
                f.get("verdict", "—"),
                "YES" if flag else "",
                "",   # Accuracy — dropdown
                "",   # Flag Rating — dropdown
                "",   # Notes — free text
            ], bg=bg, height=28)
            if flag:
                ws.cell(row=row, column=6).font = _font(bold=True, color=RED)
            # Light yellow bg on feedback cells to signal they're editable
            for col in (7, 8, 9):
                ws.cell(row=row, column=col).fill = _fill("FFFFF0")
            row += 1

    data_end_row = row - 1

    # ── Dropdowns via DataValidation ──────────────────────────────────────────
    if data_end_row >= data_start_row:
        dv_accuracy = DataValidation(
            type="list",
            formula1='"correct,partially correct,wrong,not relevant"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
        )
        dv_accuracy.sqref = f"G{data_start_row}:G{data_end_row}"

        dv_flag = DataValidation(
            type="list",
            formula1='"flag justified,over-flagged,should have been flagged,n/a"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
        )
        dv_flag.sqref = f"H{data_start_row}:H{data_end_row}"

        ws.add_data_validation(dv_accuracy)
        ws.add_data_validation(dv_flag)


# ── Main ───────────────────────────────────────────────────────────────────────

def run(company: str, output_dir: Path = None, version: str = None) -> Path:
    data    = load_data(company, version=version)
    safe    = company.replace(" ", "_").replace("/", "-")
    out_dir = output_dir or (WORKDIR / safe)
    suffix  = f"_{version}" if version else ""
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{safe}_Scorecard{suffix}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary_sheet(       wb.create_sheet("Summary"),        data)
    build_growth_metrics_sheet(wb.create_sheet("Growth Metrics"), data)
    build_tech_sheet(          wb.create_sheet("Tech Score"),     data)
    build_checklist_sheet(     wb.create_sheet("Checklist"),      data)
    build_findings_sheet(      wb.create_sheet("All Findings"),   data)

    wb.save(str(out_path))
    print(f"Scorecard: {out_path}")
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate DD scoring spreadsheet")
    parser.add_argument("company", help="Company name")
    parser.add_argument("--output", help="Output directory (default: workdir/[company])")
    args   = parser.parse_args()
    output = Path(args.output) if args.output else None
    run(args.company, output_dir=output)
