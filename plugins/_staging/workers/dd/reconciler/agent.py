"""
reconciler/agent.py — Reconciler agent.

Reads a human-reviewed scorecard (Accuracy / Flag Rating / Notes columns)
and produces corrected v2 agent JSONs. These feed into the existing
overview → appendix → format pipeline to generate v2 PDFs automatically.

Input:
    workdir/[company]/agents/*.json          (v1 pipeline outputs)
    [Company]_Reviewed_Scorecard.xlsx        (human feedback)

Output:
    workdir/[company]/agents_v2/*.json       (corrected agent outputs)
    workdir/[company]/reconciler_log.json    (record of every change made)

Correction rules:
    Accuracy = "wrong"            → replace our_finding with reviewer note;
                                    set verdict = "reconciler_corrected";
                                    remove flag if Flag Rating = "over-flagged"
    Accuracy = "partially correct"→ append reviewer note to our_finding;
                                    set verdict = "reconciler_amended"
    Flag Rating = "over-flagged"  → flag = False, flag_reason = None
    Flag Rating = "flag justified" → flag unchanged
    Accuracy = "correct"          → no change
    No feedback                   → no change
"""

import json
import copy
import re
from pathlib import Path
from datetime import datetime

import openpyxl

from config.settings import WORKDIR

AGENTS = ["financials", "comp", "qualitative", "product", "news"]

# Map scorecard Agent column values to agent JSON filenames
AGENT_NAME_MAP = {
    "financials": "financials",
    "comp":       "comp",
    "qualitative":"qualitative",
    "product":    "product",
    "news":       "news",
}

# Accuracy values that trigger content correction
WRONG_VALUES     = {"wrong"}
PARTIAL_VALUES   = {"partially correct"}
CORRECT_VALUES   = {"correct", "not relevant"}

# Flag rating values
OVER_FLAGGED     = {"over-flagged"}
JUSTIFIED        = {"flag justified"}
SHOULD_HAVE      = {"should have been flagged"}


def find_reviewed_scorecard(company: str) -> Path | None:
    """Search workdir and DD_Reports for the reviewed scorecard."""
    safe = company.replace(" ", "_").replace("/", "-")
    search_dirs = [
        WORKDIR / safe,
        Path("/mnt/c/Users/nathan/OneDrive/Desktop/CLAUDE WORK/OUTPUTS/reports/DD_Reports") / safe,
        Path("/mnt/c/Users/nathan/OneDrive/Desktop/CLAUDE WORK/OUTPUTS/reports/DD_Reports") / company,
    ]
    patterns = [
        f"Reviewed*{company}*Scorecard*.xlsx",
        f"Reviewed*{safe}*Scorecard*.xlsx",
        f"*Reviewed*Scorecard*.xlsx",
        f"*reviewed*scorecard*.xlsx",
    ]
    for d in search_dirs:
        if not d.exists():
            continue
        for pattern in patterns:
            matches = list(d.glob(pattern))
            if matches:
                return matches[0]
    return None


def load_reviewed_scorecard(path: Path) -> list[dict]:
    """
    Parse the All Findings sheet from a reviewed scorecard.
    Returns list of dicts with keys: agent, topic, our_finding, accuracy, flag_rating, notes.
    Skips rows with no feedback.
    """
    wb = openpyxl.load_workbook(str(path))

    if "All Findings" not in wb.sheetnames:
        raise ValueError(f"No 'All Findings' sheet in {path}")

    ws = wb["All Findings"]
    feedback = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        agent       = str(row[0] or "").strip().lower()
        topic       = str(row[1] or "").strip().lower().replace(" ", "_")
        our_finding = str(row[2] or "").strip()
        accuracy    = str(row[6] or "").strip().lower() if row[6] else None
        flag_rating = str(row[7] or "").strip().lower() if row[7] else None
        notes       = str(row[8] or "").strip() if row[8] else None

        # Skip rows with no feedback at all
        if not accuracy and not flag_rating and not notes:
            continue
        # Skip header rows that leaked through
        if agent in ("agent", "accuracy", ""):
            continue

        feedback.append({
            "agent":       agent,
            "topic":       topic,
            "our_finding": our_finding,
            "accuracy":    accuracy,
            "flag_rating": flag_rating,
            "notes":       notes,
        })

    return feedback


def build_feedback_index(feedback: list[dict]) -> dict:
    """Index feedback by (agent, topic) for fast lookup."""
    index = {}
    for f in feedback:
        key = (f["agent"], f["topic"])
        index[key] = f
    return index


def apply_corrections(agent_data: dict, feedback_index: dict, agent_name: str) -> tuple[dict, list]:
    """
    Apply reviewer corrections to a single agent's output.
    Returns (corrected_data, list_of_changes).
    """
    corrected = copy.deepcopy(agent_data)
    changes = []

    for i, finding in enumerate(corrected.get("findings", [])):
        topic = finding.get("topic", "").lower().replace(" ", "_")
        key   = (agent_name, topic)
        fb    = feedback_index.get(key)

        if not fb:
            continue

        accuracy    = fb.get("accuracy") or ""
        flag_rating = fb.get("flag_rating") or ""
        notes       = fb.get("notes") or ""

        original_finding  = finding.get("our_finding", "")
        original_flag     = finding.get("flag", False)
        original_verdict  = finding.get("verdict", "")

        changed = False

        # ── Content correction ────────────────────────────────────────────────
        if accuracy in WRONG_VALUES:
            if notes:
                finding["our_finding"] = f"[RECONCILER CORRECTION] {notes}"
                finding["verdict"]     = "reconciler_corrected"
                finding["confidence"]  = "high"
                finding["delta"]       = f"Original pipeline finding was incorrect. Reviewer correction: {notes}"
            else:
                finding["our_finding"] = "[RECONCILER: Finding marked wrong by reviewer — no correction note provided. Requires manual review.]"
                finding["verdict"]     = "reconciler_flagged_for_review"
            changed = True

        elif accuracy in PARTIAL_VALUES and notes:
            finding["our_finding"] = f"{original_finding}\n\n[RECONCILER AMENDMENT] {notes}"
            finding["verdict"]     = "reconciler_amended"
            finding["delta"]       = f"Pipeline finding was partially correct. Amendment: {notes}"
            changed = True

        # ── Flag correction ───────────────────────────────────────────────────
        if flag_rating in OVER_FLAGGED and finding.get("flag"):
            finding["flag"]        = False
            finding["flag_reason"] = None
            changed = True

        elif flag_rating in SHOULD_HAVE and not finding.get("flag"):
            finding["flag"]        = True
            finding["flag_reason"] = f"[RECONCILER] Reviewer indicated this should have been flagged. {notes or ''}".strip()
            changed = True

        if changed:
            changes.append({
                "finding_id":       finding.get("id", f"{agent_name}_{i}"),
                "topic":            topic,
                "accuracy":         accuracy,
                "flag_rating":      flag_rating,
                "original_finding": original_finding,
                "corrected_finding":finding["our_finding"],
                "original_flag":    original_flag,
                "corrected_flag":   finding.get("flag"),
                "original_verdict": original_verdict,
                "corrected_verdict":finding.get("verdict"),
                "reviewer_note":    notes,
            })

    # Rebuild flags list from corrected findings (maintain original list-of-dicts format)
    corrected["flags"] = [
        f for f in corrected["findings"]
        if isinstance(f, dict) and f.get("flag")
    ]

    return corrected, changes


def run(company: str, scorecard_path: Path = None) -> dict:
    """
    Run the Reconciler for a company.
    Reads reviewed scorecard, applies corrections, writes agents_v2/*.json.
    Returns a reconciler log dict.
    """
    safe        = company.replace(" ", "_").replace("/", "-")
    company_dir = WORKDIR / safe
    agents_dir  = company_dir / "agents"
    v2_dir      = company_dir / "agents_v2"
    v2_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nReconciler — {company}")
    print(f"{'='*60}")

    # ── Find and load reviewed scorecard ──────────────────────────────────────
    if scorecard_path is None:
        scorecard_path = find_reviewed_scorecard(company)

    if not scorecard_path or not scorecard_path.exists():
        raise FileNotFoundError(
            f"No reviewed scorecard found for {company}. "
            f"Expected a file matching 'Reviewed*{company}*Scorecard*.xlsx'"
        )

    print(f"Scorecard: {scorecard_path.name}")
    feedback     = load_reviewed_scorecard(scorecard_path)
    fb_index     = build_feedback_index(feedback)
    print(f"Feedback entries loaded: {len(feedback)}")

    # ── Apply corrections to each agent ───────────────────────────────────────
    all_changes  = {}
    total_changes = 0

    for agent_name in AGENTS:
        src = agents_dir / f"{agent_name}.json"
        dst = v2_dir / f"{agent_name}.json"

        if not src.exists():
            print(f"  {agent_name}: not found — skipping")
            continue

        agent_data = json.loads(src.read_text())
        corrected, changes = apply_corrections(agent_data, fb_index, agent_name)

        # Tag the output as v2
        corrected["reconciler_version"] = "v2"
        corrected["reconciler_date"]    = datetime.now().strftime("%Y-%m-%d")

        dst.write_text(json.dumps(corrected, indent=2))
        all_changes[agent_name] = changes
        total_changes += len(changes)

        print(f"  {agent_name}: {len(changes)} corrections applied → {dst.name}")

    # ── Write reconciler log ──────────────────────────────────────────────────
    log = {
        "company":          company,
        "reconciler_date":  datetime.now().isoformat(),
        "scorecard_used":   str(scorecard_path),
        "feedback_count":   len(feedback),
        "total_changes":    total_changes,
        "changes_by_agent": all_changes,
        "summary": {
            "wrong_corrected":    sum(
                1 for changes in all_changes.values()
                for c in changes if c["accuracy"] == "wrong"
            ),
            "partial_amended":    sum(
                1 for changes in all_changes.values()
                for c in changes if c["accuracy"] == "partially correct"
            ),
            "flags_removed":      sum(
                1 for changes in all_changes.values()
                for c in changes if c["original_flag"] and not c["corrected_flag"]
            ),
            "flags_added":        sum(
                1 for changes in all_changes.values()
                for c in changes if not c["original_flag"] and c["corrected_flag"]
            ),
        }
    }

    log_path = company_dir / "reconciler_log.json"
    log_path.write_text(json.dumps(log, indent=2))
    print(f"\nReconciler log: {log_path}")
    print(f"Total corrections: {total_changes}")
    print(f"  Wrong corrected:  {log['summary']['wrong_corrected']}")
    print(f"  Partial amended:  {log['summary']['partial_amended']}")
    print(f"  Flags removed:    {log['summary']['flags_removed']}")
    print(f"  Flags added:      {log['summary']['flags_added']}")

    return log
